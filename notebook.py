

'''
==================================================================================================
                Перед запуском необходимо загрузить данные библиотеки

                *   pip install --pre pycaret
                *   pip install matplotlib==3.3.0
                *   pip install pycaret[models]
                *   pip install openpyxl
==================================================================================================
'''


import numpy as np

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from datetime import date, timedelta

import os
import time
import collections.abc
collections.Iterable = collections.abc.Iterable
from pycaret.time_series import *

def prognoz(data,n_forecasts):
    ret = []
    for i in data:
        df = data.copy()
        df.drop(df.tail(n_forecasts).index,inplace=True)
        #print(df.tail())

        y = df[i].to_numpy()
        y = y[n_forecasts:]
        df_exog = df[-n_forecasts:]
        df_exog =  df_exog.drop(columns = i,axis = 1)
        df = df[:(df[i].count() - n_forecasts)]
        df[i] = y

        ''' MODEL '''
        pred = predicted(df,df_exog,n_forecasts,str(i))
        ret.append(pred)
    return(ret)

def prognoz_hold(data, forecastes):
    ret = []
    #print(forecastes)
    max_for = np.amax(forecastes)
    for i in data:
        df = data.copy()
        index_for = df.columns.get_loc(str(i))
        n_forecasts = forecastes[index_for]
        if forecastes[index_for] < max_for:
            last_index = df[i].count()-forecastes[index_for]
            df_srez = df[:last_index]
            mas_del = []
            for j in df_srez:
                if 'Forecast' in df_srez[j].to_numpy():
                    mas_del.append(j)
            df_srez.drop(mas_del, axis=1, inplace=True)

            y = df_srez[i].to_numpy()
            y = y[n_forecasts:]
            df_exog = df_srez[-n_forecasts:]
            df_exog =  df_exog.drop(columns = i,axis = 1)
            df_srez = df_srez[:(df_srez[i].count() - n_forecasts)]
            df_srez[i] = y

            ''' MODEL '''
            pred = predicted(df_srez,df_exog,n_forecasts,str(i))
            ret.append(pred)
        else:
            last_index = df[i].count()-forecastes[index_for]
            df_srez = df[:last_index]

            y = df_srez[i].to_numpy()
            y = y[n_forecasts:]
            df_exog = df_srez[-n_forecasts:]
            df_exog =  df_exog.drop(columns = i,axis = 1)
            df_srez = df_srez[:(df_srez[i].count() - n_forecasts)]
            df_srez[i] = y

            ''' MODEL '''
            pred = predicted(df_srez,df_exog,n_forecasts,str(i))
            ret.append(pred)
    return(ret)

def prognoz_if(data,forecastes):
    ret = []
    #print(forecastes)
    df_exog = data.tail(np.amax(forecastes))
    mas_del = []
    for i in range(len(forecastes)):
        if forecastes[i] > 0:
            mas_del.append(i)
    df_exog = df_exog.drop(df_exog.columns[mas_del], axis=1)
    for i in data:
        if 'Forecast' in data[i].to_numpy():
            df = data.copy()
            index_for = df.columns.get_loc(str(i))
            n_forecasts = forecastes[index_for]
            df.drop(df.tail(np.amax(forecastes)).index,inplace=True)

            y = df[i].to_numpy()
            list_exog = list(df_exog)
            for j in df:
                if not(j in list_exog):
                    df =  df.drop(columns = j,axis = 1)
            #print(df)
            df[i] = y

            ''' MODEL '''
            pred = predicted(df,df_exog,n_forecasts,str(i))
            ret.append(pred)
    return(ret)

def prognoz_if_hold(data,forecastes):
    ret = []
    for i in data:
        if 'Forecast' in data[i].to_numpy():
            df = data.copy()
            index_for = df.columns.get_loc(str(i))
            n_forecasts = forecastes[index_for]
            df_exog = data.tail(np.amax(n_forecasts))
            mas_del = []
            for k in range(len(forecastes)):
                if forecastes[k] > 0:
                    mas_del.append(k)
            df_exog = df_exog.drop(df_exog.columns[mas_del], axis=1)
            #print(df_exog)
            df_srez = pd.DataFrame()
            for j in df_exog:
                df_srez[j] = df[j][:(df[j].count() - n_forecasts)]
            df_srez[i] = df[i][:(df[i].count() - n_forecasts)]


            ''' MODEL '''
            pred = predicted(df_srez,df_exog,n_forecasts,str(i))
            ret.append(pred)
    return(ret)



def predicted(data, df_exog, forecasting, colum, flag=False):

    data = data.astype(np.float)
    df_exog = df_exog.astype(np.float)
    #print(df_exog.empty)

    if int(forecasting)!=0:
        if flag == False:
            #data = data.astype(np.float)
            s = setup(data, target = colum, fh = int(forecasting), enforce_exogenous = not(df_exog.empty), verbose=False) #, use_gpu = True
            models  = create_model('lightgbm_cds_dt', verbose=False)

        elif flag == True:
            #df = df.astype(np.float)
            s = setup(data, target = colum, fh = int(forecasting), enforce_exogenous = not(df_exog.empty), verbose=False) #, use_gpu = True
            models = compare_models(verbose=False)

        test_data = data.drop(colum, axis = 1)
        predictions_future = predict_model(models, X = test_data)
        predictions_future = predictions_future['y_pred'].to_numpy()

        return predictions_future
    else:
        return None

def type_task(data):

    forecastes = np.array([])

    #data = data.set_index("DATE")
    for i in data:
        try:
            count = data[str(i)].value_counts()["Forecast"]
            forecastes = np.append(forecastes, count)
        except:
            forecastes = np.append(forecastes, 0)

    forecastes = forecastes.astype(int)

    if np.unique(forecastes).size == 1:
        #print("Прогноз")
        otvet = prognoz(data,forecastes[0])
    else:
        if 0 in forecastes:
            if np.unique(forecastes).size == 2:
                #print("Условный прогноз")
                otvet = prognoz_if(data,forecastes)
            else:
                #print("Условная задержка")
                otvet = prognoz_if_hold(data,forecastes)
        else:
            #print("Задержка")
            otvet = prognoz_hold(data,forecastes)
    return otvet

start_time = time.time()


def data_exel_read_write(filename_read, filename_save):

    exel = load_workbook(filename_read) #загружаем ексель файл как объект
    sheet = exel.sheetnames #получаем в виде списка названия листов
    wb = Workbook()

    def datatime_func(Data,name):

        if name=='Monthly' or 'Quarterly':
            data = Data
            a = np.array([])
            data = data.rename(columns={None : 'DATE'})
            datatime = data['DATE'].to_numpy()
            d1 = date(2000 + int(datatime[0][:2]), int(datatime[0][-2:]), 1)  # начальная дата 2001-01-01
            d2 = date(2000 + int(datatime[len(datatime)-1][:2]), int(datatime[len(datatime)-1][-2:]), 1)  # конечная дат
            #print(type(d1))
            #print(type(d2))
            delta = d2 - d1         # timedelta
            for i in range(delta.days + 1):
                if (d1 + timedelta(i)).day == 1:
                    a = np.append(a, (d1 + timedelta(i)))
            data['DATE'] = a[:len(datatime)]
            data["DATE"] = pd.to_datetime(data["DATE"])
            data = data.set_index("DATE")
            #data.asfreq('M')
            #print(type(df))
            return datatime, data


    for name in sheet:
        ws = exel[name]
        data = ws.values
        columns = next(data)[0:]
        Data = pd.DataFrame(data, columns=columns)
        datatime, data = datatime_func(Data,name) #получаем исправленный датасет с нормальным временем

        #network(data)
        #prophet(data)


        '''Меняю данные'''

        #пример данных
        #print(data)
        array = type_task(data)
        data.insert(0, '', datatime)
        #array = np.zeros((4, 12)) #ответы


        #print(data)
        #print(array)
        #print(filename_save)
        minus_enum = 0
        for element in enumerate(list(data)[1:]):
            if ('Forecast' in data[element[1]].to_numpy()):
                columns = data[data[element[1]]!='Forecast'][element[1]].to_numpy() #Берём колонку без форекастов
                columns = np.append(columns, array[element[0]-minus_enum]) #добовляем значения из ответов
                data[element[1]] = columns # присваиваем
            else:
                minus_enum+=1


        '''------------'''

        '''Запись на новый лиист значений'''
        wb.create_sheet(name)
        ws = wb[name]
        for r in dataframe_to_rows(data, index=False, header=True): #Запись построково значений
            ws.append(r)
        #break

    name_list = wb.sheetnames
    del wb[name_list[0]]
    wb.save(filename_save)
    wb.close()




#data_exel_read_write('Тестовый датасет/Test_input_5.xlsx', "Тестовый датасет/Test_output_5.xlsx")


folder_dir = 'Тут датасет/Тестовый датасет'
folder_dir_answer = 'Тут датасет/Ответы'
for file in os.listdir(folder_dir):
    try:
        data_exel_read_write((folder_dir + '/' + file), (folder_dir_answer + '/' + file.replace('input', 'output')))
    except:
        print(file)
    #break



print("--- %s Время: ---" % (time.time() - start_time))